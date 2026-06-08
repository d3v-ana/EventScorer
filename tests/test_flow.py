"""全流程测试 — 覆盖完整的业务链路"""

import io
import json
from openpyxl import load_workbook


class TestFullFlow:
    """完整的端到端业务流程测试"""

    def test_admin_login_flow(self, client):
        """管理员登录 → 首页"""
        resp = client.get('/login')
        assert resp.status_code == 200

        resp = client.post('/login', data={
            'email': 'tenant@example.com', 'password': 'admin',
        }, follow_redirects=True)
        assert resp.status_code == 200
        html = resp.data.decode('utf-8')
        assert '活动' in html

    def test_create_department(self, auth_client):
        """创建部门"""
        resp = auth_client.post('/admin/department/create', data={
            'name': '体育部', 'sort_order': 1,
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert '体育部' in resp.data.decode('utf-8')

    def test_full_business_flow(self, auth_client, app):
        """完整业务流程：创建项目→活动→二维码→录入员→注册→录入→排名"""
        from app import db
        from app.models import (
            Activity, ActivityProject, ActivityType, Department,
            Project, QRCode, Recorder, ActivityRecorder,
            Participant, Result,
        )

        # ---- 1. 使用已有的默认活动类型和部门 ----
        at = ActivityType.query.filter_by(name='学生').first()
        assert at is not None, '默认活动类型 "学生" 不存在'
        dept = Department.query.filter_by(name='体育部').first()
        if not dept:
            dept = Department(name='体育部', sort_order=1)
            db.session.add(dept)
            db.session.flush()

        # ---- 2. 创建时间型项目 ----
        resp = auth_client.post('/admin/project/create', data={
            'name': '100米短跑',
            'type': 'time',
            'penalty': 5.0,
            'category_id': 0,
        }, follow_redirects=True)
        assert resp.status_code == 200
        proj_time = Project.query.filter_by(name='100米短跑').first()
        assert proj_time is not None
        assert proj_time.type == 'time'

        # ---- 3. 创建分数型项目 ----
        resp = auth_client.post('/admin/project/create', data={
            'name': '引体向上',
            'type': 'score',
            'max_score': 100,
        }, follow_redirects=True)
        assert resp.status_code == 200
        proj_score = Project.query.filter_by(name='引体向上').first()
        assert proj_score is not None
        assert proj_score.type == 'score'
        assert proj_score.max_score == 100

        # ---- 4. 创建活动 ----
        resp = auth_client.post('/admin/activity/create', data={
            'name': '2024年秋季运动会',
            'activity_type_id': at.id,
            'department_id': dept.id,
            'project_ids': [str(proj_time.id), str(proj_score.id)],
        }, follow_redirects=True)
        assert resp.status_code == 200
        activity = Activity.query.filter_by(name='2024年秋季运动会').first()
        assert activity is not None
        aps = ActivityProject.query.filter_by(activity_id=activity.id).all()
        assert len(aps) == 2

        # ---- 4.5 配置自定义词库 ----
        activity.custom_words = '\n'.join(['红狐','蓝鲸','金鹰','银狼'])
        db.session.commit()

        # ---- 5. 生成二维码 ----
        resp = auth_client.post(f'/admin/activity/{activity.id}/qrcode/generate', data={
            'count': 5,
        })
        data = json.loads(resp.data)
        assert resp.status_code == 200
        assert 'qrcodes' in data
        assert len(data['qrcodes']) == 5
        qrs = QRCode.query.filter_by(activity_id=activity.id).all()
        assert len(qrs) == 5
        qr = qrs[0]

        # ---- 6. 创建录入员 ----
        resp = auth_client.post('/admin/recorder/create', data={
            'name': '录入员张三',
        }, follow_redirects=True)
        assert resp.status_code == 200
        recorder = Recorder.query.filter_by(name='录入员张三').first()
        assert recorder is not None
        assert len(recorder.record_key) == 4

        # ---- 7. 分配录入员到活动（含项目权限） ----
        resp = auth_client.post(f'/admin/activity/{activity.id}/recorder/assign', data={
            'recorder_id': recorder.id,
            'project_ids': [str(proj_time.id), str(proj_score.id)],
        }, follow_redirects=True)
        assert resp.status_code == 200
        ar = ActivityRecorder.query.filter_by(
            activity_id=activity.id, recorder_id=recorder.id
        ).first()
        assert ar is not None
        assert proj_time.id in ar.project_id_list()
        assert proj_score.id in ar.project_id_list()

        # ---- 8. 参与者通过二维码注册 ----
        resp = auth_client.post(f'/register/{qr.code}', data={
            'name': '测试学生甲',
            'class_name': '高一(1)班',
        }, follow_redirects=True)
        assert resp.status_code == 200
        participant = Participant.query.filter_by(name='测试学生甲').first()
        assert participant is not None
        assert participant.class_name == '高一(1)班'
        assert QRCode.query.get(qr.id).status == 'used'

        # 设置录入员 session
        with auth_client.session_transaction() as sess:
            sess['recorder_id'] = recorder.id
            sess['recorder_key'] = recorder.record_key
            sess['activity_id'] = activity.id

        # ---- 9. 录入时间型成绩 ----
        resp = auth_client.post(
            f'/recorder/{recorder.id}/input/{qr.code}',
            data={
                'project_id': proj_time.id,
                'time_minutes': 1,
                'time_seconds': 30,
                'time_ms': 500,
                'violations': 2,
            },
        )
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['success'] is True
        assert data['type'] == 'time'
        # 1:30.500 + 2*5 = 90.5 + 10 = 100.5
        assert abs(data['final_time'] - 100.5) < 0.01

        # ---- 10. 录入分数型成绩 ----
        resp = auth_client.post(
            f'/recorder/{recorder.id}/input/{qr.code}',
            data={
                'project_id': proj_score.id,
                'score': 85,
            },
        )
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['success'] is True
        assert data['type'] == 'score'
        assert abs(data['final_time'] - 85) < 0.01

        # ---- 11. 验证结果 ----
        results = Result.query.filter_by(participant_id=participant.id).all()
        assert len(results) == 2
        r_time = next(r for r in results if r.project_id == proj_time.id)
        assert abs(r_time.final_time - 100.5) < 0.01
        assert r_time.violations == 2
        r_score = next(r for r in results if r.project_id == proj_score.id)
        assert abs(r_score.final_time - 85) < 0.01

        # ---- 12. 查看排名 ----
        resp = auth_client.get(f'/admin/activity/{activity.id}/ranking')
        assert resp.status_code == 200
        html = resp.data.decode('utf-8')
        assert '测试学生甲' in html

        # ---- 13. 导出 Excel ----
        resp = auth_client.get(f'/admin/activity/{activity.id}/export')
        assert resp.status_code == 200
        assert 'spreadsheetml' in resp.content_type or 'officedocument' in resp.content_type
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        assert ws.title == '2024年秋季运动会'
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert '100米短跑(用时)' in headers
        assert '引体向上(成绩)' in headers

        # ---- 14. 参与者查看成绩 ----
        resp = auth_client.get(f'/user/results/{qr.code}')
        assert resp.status_code == 200
        html = resp.data.decode('utf-8')
        assert '测试学生甲' in html

    def test_flow_multi_participant(self, auth_client, app):
        """多参与者排名测试"""
        from app import db
        from app.models import (
            Activity, ActivityProject, ActivityType,
            Project, QRCode, Recorder, ActivityRecorder, Participant, Result,
        )
        at = ActivityType.query.filter_by(name='学生').first()
        assert at is not None

        p1 = Project(name='50米', type='time', penalty_per_violation=5.0)
        db.session.add(p1)
        p2 = Project(name='立定跳远', type='time', penalty_per_violation=2.0)
        db.session.add(p2)
        db.session.flush()

        act = Activity(name='体能测试', activity_type_id=at.id)
        db.session.add(act)
        db.session.flush()
        db.session.add(ActivityProject(activity_id=act.id, project_id=p1.id))
        db.session.add(ActivityProject(activity_id=act.id, project_id=p2.id))

        rec = Recorder(name='录入员李四', record_key='0002')
        db.session.add(rec)
        db.session.flush()
        db.session.add(
            ActivityRecorder(activity_id=act.id, recorder_id=rec.id,
                             project_ids=f'{p1.id},{p2.id}')
        )

        participant_qrs = []
        for name, cls in [('甲同学', '一班'), ('乙同学', '一班'), ('丙同学', '二班')]:
            qr = QRCode(code=f'T{len(participant_qrs)+1:04d}',
                        activity_id=act.id, status='used')
            db.session.add(qr)
            db.session.flush()
            p = Participant(name=name, class_name=cls,
                            activity_id=act.id, qrcode_id=qr.id)
            db.session.add(p)
            db.session.flush()
            participant_qrs.append((p, qr.code))
        db.session.commit()

        # 录入员 session
        with auth_client.session_transaction() as sess:
            sess['recorder_id'] = rec.id
            sess['recorder_key'] = rec.record_key
            sess['activity_id'] = act.id

        # 录入成绩
        scores = [
            (participant_qrs[0], 10, 5),   # 甲: 15s
            (participant_qrs[1], 12, 8),   # 乙: 20s
            (participant_qrs[2], 15, 10),  # 丙: 25s
        ]
        for (p, p_qr), s1, s2 in scores:
            auth_client.post(f'/recorder/{rec.id}/input/{p_qr}', data={
                'project_id': p1.id, 'time_minutes': 0, 'time_seconds': s1,
                'time_ms': 0, 'violations': 0,
            })
            auth_client.post(f'/recorder/{rec.id}/input/{p_qr}', data={
                'project_id': p2.id, 'time_minutes': 0, 'time_seconds': s2,
                'time_ms': 0, 'violations': 0,
            })

        resp = auth_client.get(f'/admin/activity/{act.id}/ranking')
        assert resp.status_code == 200
        html = resp.data.decode('utf-8')
        assert '甲同学' in html
        assert '乙同学' in html
        assert '丙同学' in html

    def test_activity_archive_restore_delete(self, auth_client, app):
        """活动归档、恢复、删除（不依赖 activity_detail 模板）"""
        from app import db
        from app.models import Activity, ActivityType, Project, ActivityProject
        at = ActivityType.query.filter_by(name='学生').first()
        assert at is not None
        p = Project(name='铅球', type='score', max_score=100)
        db.session.add(p)
        db.session.flush()
        act = Activity(name='教职工运动会', activity_type_id=at.id)
        db.session.add(act)
        db.session.flush()
        db.session.add(ActivityProject(activity_id=act.id, project_id=p.id))
        db.session.commit()
        aid = act.id

        # 归档 → 跳转到活动列表页
        resp = auth_client.post(f'/admin/activity/{aid}/archive', follow_redirects=True)
        assert resp.status_code == 200
        assert Activity.query.get(aid).archived is True

        # 恢复
        resp = auth_client.post(f'/admin/activity/{aid}/unarchive', follow_redirects=True)
        assert resp.status_code == 200
        assert Activity.query.get(aid).archived is False

        # 删除
        resp = auth_client.post(f'/admin/activity/{aid}/delete', follow_redirects=True)
        assert resp.status_code == 200
        assert Activity.query.get(aid) is None
