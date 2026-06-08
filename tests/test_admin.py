"""Tests for admin routes."""
import io
import json
from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook

from app import db
from app.models import (
    Admin, Activity, ActivityProject, ActivityType, Department, Participant,
    Project, ProjectCategory, QRCode, Recorder, Result
)


class TestAdminOverview:
    def test_admin_page(self, auth_client):
        resp = auth_client.get('/admin')
        assert resp.status_code == 200

    def test_common_admin_buttons_use_icons(
            self, auth_client, sample_activity, sample_project,
            sample_recorder):
        admin_body = auth_client.get('/admin').data.decode('utf-8')
        project_body = auth_client.get('/admin/projects').data.decode('utf-8')
        recorder_body = auth_client.get('/admin/recorders').data.decode('utf-8')

        assert '🔍 搜索' in admin_body
        assert '📋 详情' in admin_body
        assert '🗄️ 归档' in admin_body
        assert '✏️ 编辑' in project_body
        assert '🗑️ 删除' in project_body
        assert '➕ 创建录入员' in recorder_body

    def test_activity_created_at_shows_date_only(
            self, auth_client, app, sample_activity):
        with app.app_context():
            sample_activity.created_at = datetime(
                2026, 5, 30, 14, 25, tzinfo=timezone.utc
            )
            db.session.commit()

        resp = auth_client.get('/admin')
        body = resp.data.decode('utf-8')

        assert '2026-05-30' in body
        assert '2026-05-30 14:25' not in body

    def test_archived_activity_created_at_shows_date_only(
            self, auth_client, app, sample_activity):
        with app.app_context():
            sample_activity.archived = True
            sample_activity.created_at = datetime(
                2026, 5, 30, 14, 25, tzinfo=timezone.utc
            )
            db.session.commit()

        resp = auth_client.get('/admin/archived')
        body = resp.data.decode('utf-8')

        assert '2026-05-30' in body
        assert '2026-05-30 14:25' not in body

    def test_other_admin_created_at_columns_show_date_only(
            self, auth_client, app, sample_project):
        stamp = datetime(2026, 5, 30, 14, 25, tzinfo=timezone.utc)
        with app.app_context():
            sample_project.created_at = stamp
            category = ProjectCategory(name='Date Only Group',
                                       created_at=stamp)
            department = Department(name='Date Only Department',
                                    created_at=stamp)
            activity_type = ActivityType(name='Date Only Type',
                                         created_at=stamp)
            admin = Admin(email='dateonly@example.com', username='dateonly',
                          role='tenant_admin',
                          password_hash='unused',
                          created_at=stamp)
            db.session.add_all([category, department, activity_type, admin])
            db.session.commit()

        paths = [
            '/admin/projects',
            '/admin/categories',
            '/admin/departments',
            '/admin/activity-types',
            '/admin/account',
        ]
        for path in paths:
            body = auth_client.get(path).data.decode('utf-8')
            assert '2026-05-30' in body
            assert '2026-05-30 14:25' not in body

    def test_system_logs_has_mobile_responsive_table_wrapper(
            self, auth_client):
        resp = auth_client.get('/admin/logs')
        body = resp.data.decode('utf-8')

        assert 'admin-log-table-wrap' in body
        assert 'admin-log-table' in body

    def test_create_admin_redirects_to_refreshed_account_page(
            self, auth_client):
        resp = auth_client.post('/admin/account', data={
            'action': 'create_admin',
            'username': 'freshadmin',
            'email': 'freshadmin@example.com',
            'password': 'secret123',
            '_csrf_token': 'test',
        })

        assert resp.status_code == 302
        assert resp.headers['Location'].endswith('/admin/account')

        refreshed = auth_client.get(resp.headers['Location'])
        body = refreshed.data.decode('utf-8')
        assert 'freshadmin' in body


class TestProjectCRUD:
    def test_project_list(self, auth_client):
        resp = auth_client.get('/admin/projects')
        assert resp.status_code == 200

    def test_project_pages_use_group_wording(self, auth_client):
        list_resp = auth_client.get('/admin/projects')
        group_resp = auth_client.get('/admin/categories')
        create_resp = auth_client.get('/admin/project/create')

        for resp in (list_resp, group_resp, create_resp):
            body = resp.data.decode('utf-8')
            assert '项目分组' in body or '分组' in body
            assert '项目分类' not in body

    def test_create_project_page(self, auth_client):
        resp = auth_client.get('/admin/project/create')
        assert resp.status_code == 200

    def test_project_form_labels_score_limit_not_penalty_for_score_ui(
            self, auth_client):
        resp = auth_client.get('/admin/project/create')
        body = resp.data.decode('utf-8')

        assert '最大分值' in body
        assert '犯规罚时(秒/次)' in body
        assert "maxScoreGroup').style.display=(type==='score')?'':'none'" in body
        assert "penaltyGroup').style.display=(type==='score')?'none':''" in body

    def test_create_project(self, auth_client):
        resp = auth_client.post('/admin/project/create', data={
            'name': '新项目',
            'penalty': '10.0',
            '_csrf_token': 'test'
        }, follow_redirects=True)
        assert resp.status_code == 200

    def test_edit_project_page(self, auth_client, sample_project):
        resp = auth_client.get(f'/admin/project/{sample_project.id}/edit')
        assert resp.status_code == 200

    def test_project_list_has_mobile_responsive_table_wrapper(
            self, auth_client, sample_project):
        resp = auth_client.get('/admin/projects')
        body = resp.data.decode('utf-8')

        assert 'project-list-toolbar' in body
        assert 'project-list-table-wrap' in body
        assert 'project-list-table' in body
        assert '@media (max-width: 768px)' in body

    def test_category_list_has_mobile_responsive_table_wrapper(
            self, auth_client, app):
        with app.app_context():
            db.session.add(ProjectCategory(name='Responsive Group'))
            db.session.commit()

        resp = auth_client.get('/admin/categories')
        body = resp.data.decode('utf-8')

        assert 'category-list-table-wrap' in body
        assert 'category-list-table' in body


class TestActivityCRUD:
    def test_create_activity_page(self, auth_client, sample_project):
        resp = auth_client.get('/admin/activity/create')
        assert resp.status_code == 200

    def test_create_activity_filters_projects_by_group(
            self, auth_client, app, sample_project):
        with app.app_context():
            category = ProjectCategory(name='Track Group')
            sample_project.category_id = category.id
            db.session.add(category)
            db.session.flush()
            sample_project.category_id = category.id
            db.session.commit()

        resp = auth_client.get('/admin/activity/create')
        body = resp.data.decode('utf-8')

        assert 'id="projectCategoryFilter"' in body
        assert 'data-project-filter-item' in body
        assert 'data-category-id="' in body
        assert 'Track Group' in body
        assert 'filterProjectChoices' in body

    def test_create_activity(self, auth_client, sample_project):
        resp = auth_client.post('/admin/activity/create', data={
            'name': '新活动',
            'activity_type_id': '1',
            'project_ids': [str(sample_project.id)],
            '_csrf_token': 'test'
        }, follow_redirects=True)
        assert resp.status_code == 200

    def test_activity_detail(self, auth_client, sample_activity):
        resp = auth_client.get(f'/admin/activity/{sample_activity.id}')
        assert resp.status_code == 200

    def test_activity_detail_filters_add_project_select_by_group(
            self, auth_client, app, sample_activity):
        with app.app_context():
            category = ProjectCategory(name='Field Group')
            db.session.add(category)
            db.session.flush()
            project = Project(name='Field Project', category_id=category.id)
            db.session.add(project)
            db.session.commit()

        resp = auth_client.get(f'/admin/activity/{sample_activity.id}')
        body = resp.data.decode('utf-8')

        assert 'id="addProjectCategoryFilter"' in body
        assert 'id="addProjectSelect"' in body
        assert 'data-category-id="' in body
        assert 'Field Group' in body
        assert 'filterAddProjectOptions' in body

    def test_participant_delete_confirm_mentions_results(
            self, auth_client, app, sample_activity, sample_project):
        with app.app_context():
            qr = QRCode(code='DELETE-CONFIRM',
                        activity_id=sample_activity.id,
                        status='used')
            db.session.add(qr)
            db.session.flush()
            participant = Participant(name='Delete Confirm User',
                                      activity_id=sample_activity.id,
                                      qrcode_id=qr.id)
            db.session.add(participant)
            db.session.flush()
            db.session.add(Result(participant_id=participant.id,
                                  project_id=sample_project.id,
                                  time_seconds=12,
                                  final_time=12))
            db.session.commit()

        resp = auth_client.get(f'/admin/activity/{sample_activity.id}?tab=tab-participants')
        body = resp.data.decode('utf-8')

        confirm_text = '确定删除参赛者「Delete Confirm User」及其已录入成绩吗？'
        assert f'confirm({json.dumps(confirm_text)})' in body

    def test_activity_ranking(self, auth_client, sample_activity):
        resp = auth_client.get(f'/admin/activity/{sample_activity.id}/ranking')
        assert resp.status_code == 200

    def test_activity_ranking_supports_auto_refresh(
            self, auth_client, sample_activity):
        resp = auth_client.get(f'/admin/activity/{sample_activity.id}/ranking')
        body = resp.data.decode('utf-8')

        assert 'id="rankingAutoRefreshRoot"' in body
        assert 'data-refresh-interval="10000"' in body
        assert 'setInterval(refreshRankings' in body

    def test_activity_ranking_shows_mixed_time_and_score_totals(
            self, auth_client, app, sample_activity, sample_project):
        with app.app_context():
            score_project = Project(name='Score Project', type='score',
                                    max_score=100)
            db.session.add(score_project)
            db.session.flush()
            db.session.add(ActivityProject(activity_id=sample_activity.id,
                                           project_id=score_project.id))

            qr_fast = QRCode(code='RANK-MIXED-FAST',
                             activity_id=sample_activity.id)
            qr_slow = QRCode(code='RANK-MIXED-SLOW',
                             activity_id=sample_activity.id)
            db.session.add_all([qr_fast, qr_slow])
            db.session.flush()
            fast = Participant(name='Fast Low Score',
                               activity_id=sample_activity.id,
                               qrcode_id=qr_fast.id)
            slow = Participant(name='Slow High Score',
                               activity_id=sample_activity.id,
                               qrcode_id=qr_slow.id)
            db.session.add_all([fast, slow])
            db.session.flush()
            db.session.add_all([
                Result(participant_id=fast.id, project_id=sample_project.id,
                       time_seconds=10, final_time=10),
                Result(participant_id=fast.id, project_id=score_project.id,
                       time_seconds=50, final_time=50),
                Result(participant_id=slow.id, project_id=sample_project.id,
                       time_seconds=20, final_time=20),
                Result(participant_id=slow.id, project_id=score_project.id,
                       time_seconds=90, final_time=90),
            ])
            db.session.commit()

        resp = auth_client.get(f'/admin/activity/{sample_activity.id}/ranking')
        body = resp.data.decode('utf-8')

        assert '总用时' in body
        assert '总分' in body
        assert '最高总分' in body
        assert '0:10.000' in body
        assert '0:20.000' in body
        assert '90' in body

    def test_activity_detail_displays_configured_activity_type(
            self, auth_client, app):
        with app.app_context():
            activity_type = ActivityType(name='亲子组', sort_order=0)
            db.session.add(activity_type)
            db.session.flush()
            activity = Activity(name='亲子活动',
                                activity_type_id=activity_type.id)
            db.session.add(activity)
            db.session.commit()
            activity_id = activity.id

        resp = auth_client.get(f'/admin/activity/{activity_id}')

        body = resp.data.decode('utf-8')
        assert '亲子组' in body
        assert '教职工' not in body

    def test_activity_type_field_options_use_newline_textarea(
            self, auth_client):
        resp = auth_client.get('/admin/activity-types')
        body = resp.data.decode('utf-8')

        assert '<textarea' in body
        assert "join('\\n')" in body
        assert 'split(/\\r?\\n/)' in body
        assert "split(',')" not in body

    def test_activity_type_list_has_mobile_responsive_table_wrapper(
            self, auth_client):
        resp = auth_client.get('/admin/activity-types')
        body = resp.data.decode('utf-8')

        assert 'activity-type-list-table-wrap' in body
        assert 'activity-type-list-table' in body

    def test_save_activity_type_fields_accepts_line_options(
            self, auth_client, app):
        with app.app_context():
            activity_type = ActivityType.query.first()
            type_id = activity_type.id

        fields = [
            {'key': 'name', 'label': 'Name', 'type': 'text',
             'required': True},
            {'key': 'team', 'label': 'Team', 'type': 'select',
             'required': True, 'options': ['Red', 'Blue']},
        ]
        resp = auth_client.post(
            f'/admin/activity-type/{type_id}/fields',
            data={
                'fields_config': json.dumps(fields),
                '_csrf_token': 'test',
            },
        )

        assert resp.status_code == 302
        with app.app_context():
            saved = db.session.get(ActivityType, type_id).get_parsed_fields()
        assert saved[1]['options'] == ['Red', 'Blue']

    @pytest.mark.parametrize('options', [[], ['Red,Blue'], ['Red', 'Red']])
    def test_save_activity_type_fields_rejects_bad_select_options(
            self, auth_client, app, options):
        with app.app_context():
            activity_type = ActivityType.query.first()
            type_id = activity_type.id
            original = activity_type.fields_config

        fields = [
            {'key': 'name', 'label': 'Name', 'type': 'text',
             'required': True},
            {'key': 'team', 'label': 'Team', 'type': 'select',
             'required': True, 'options': options},
        ]
        resp = auth_client.post(
            f'/admin/activity-type/{type_id}/fields',
            data={
                'fields_config': json.dumps(fields),
                '_csrf_token': 'test',
            },
        )

        assert resp.status_code == 400
        assert '字段配置' in resp.data.decode('utf-8')
        with app.app_context():
            saved = db.session.get(ActivityType, type_id)
            assert saved.fields_config == original


class TestRecorderManagement:
    def test_recorder_list(self, auth_client):
        resp = auth_client.get('/admin/recorders')
        assert resp.status_code == 200

    def test_recorder_list_has_mobile_responsive_table_wrapper(
            self, auth_client, sample_recorder):
        resp = auth_client.get('/admin/recorders')
        body = resp.data.decode('utf-8')

        assert 'recorder-list-table-wrap' in body
        assert 'recorder-list-table' in body

    def test_create_recorder(self, auth_client):
        resp = auth_client.post(
            '/admin/recorder/create',
            data={
                'name': '全局录入员',
                '_csrf_token': 'test'
            },
            follow_redirects=True
        )
        assert resp.status_code == 200

    def test_edit_recorder(self, auth_client, sample_recorder):
        resp = auth_client.post(
            f'/admin/recorder/{sample_recorder.id}/edit',
            data={
                'name': '新名字',
                'record_key': '8888',
                '_csrf_token': 'test'
            },
            follow_redirects=True
        )
        assert resp.status_code == 200

    def test_delete_recorder(self, auth_client, sample_recorder):
        resp = auth_client.post(
            f'/admin/recorder/{sample_recorder.id}/delete',
            data={'_csrf_token': 'test'},
            follow_redirects=True
        )
        assert resp.status_code == 200

    def test_delete_recorder_keeps_results_with_null_recorder(
            self, auth_client, app, sample_recorder, sample_activity,
            sample_project):
        with app.app_context():
            qr = QRCode(code='RECORDER-DELETE',
                        activity_id=sample_activity.id)
            db.session.add(qr)
            db.session.flush()
            participant = Participant(name='Recorded User',
                                      activity_id=sample_activity.id,
                                      qrcode_id=qr.id)
            db.session.add(participant)
            db.session.flush()
            result = Result(participant_id=participant.id,
                            project_id=sample_project.id,
                            time_seconds=10,
                            final_time=10,
                            recorder_id=sample_recorder.id)
            db.session.add(result)
            db.session.commit()
            result_id = result.id

        resp = auth_client.post(
            f'/admin/recorder/{sample_recorder.id}/delete',
            data={'_csrf_token': 'test'},
            follow_redirects=True
        )

        assert resp.status_code == 200
        with app.app_context():
            result = db.session.get(Result, result_id)
            assert result is not None
            assert result.recorder_id is None

    def test_assign_recorder_to_activity(self, auth_client, sample_activity,
                                          sample_recorder, sample_project):
        resp = auth_client.post(
            f'/admin/activity/{sample_activity.id}/recorder/assign',
            data={
                'recorder_id': str(sample_recorder.id),
                'project_ids': [str(sample_project.id)],
                '_csrf_token': 'test'
            },
            follow_redirects=True
        )
        assert resp.status_code == 200

    def test_remove_recorder_from_activity(self, auth_client, sample_activity,
                                            sample_activity_recorder):
        resp = auth_client.post(
            f'/admin/activity-recorder/{sample_activity_recorder.id}/remove',
            data={'_csrf_token': 'test'},
            follow_redirects=True
        )
        assert resp.status_code == 200


class TestQRCodeManagement:
    def test_generate_qrcodes(self, auth_client, sample_activity, app):
        from app import db
        sample_activity.custom_words = '\n'.join(['红狐','蓝鲸','金鹰'])
        db.session.commit()
        resp = auth_client.post(
            f'/admin/activity/{sample_activity.id}/qrcode/generate',
            data={'count': '3', '_csrf_token': 'test'}
        )
        assert resp.status_code == 200
        import json
        data = json.loads(resp.data)
        assert len(data['qrcodes']) == 3

    def test_qr_print(self, auth_client, sample_activity):
        resp = auth_client.get(f'/admin/activity/{sample_activity.id}/qrprint')
        assert resp.status_code == 200


class TestExport:
    def test_export_results(self, auth_client, sample_activity):
        resp = auth_client.get(f'/admin/activity/{sample_activity.id}/export')
        assert resp.status_code == 200
        assert resp.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def test_export_mixed_activity_ranks_by_time_before_score(
            self, auth_client, app, sample_activity, sample_project):
        with app.app_context():
            score_project = Project(name='Score Project', type='score',
                                    max_score=100)
            db.session.add(score_project)
            db.session.flush()
            db.session.add(ActivityProject(activity_id=sample_activity.id,
                                           project_id=score_project.id))

            qr_fast = QRCode(code='MIXED-FAST',
                             activity_id=sample_activity.id)
            qr_slow = QRCode(code='MIXED-SLOW',
                             activity_id=sample_activity.id)
            db.session.add_all([qr_fast, qr_slow])
            db.session.flush()
            fast = Participant(name='Fast High Score',
                               activity_id=sample_activity.id,
                               qrcode_id=qr_fast.id)
            slow = Participant(name='Slow Low Score',
                               activity_id=sample_activity.id,
                               qrcode_id=qr_slow.id)
            db.session.add_all([fast, slow])
            db.session.flush()
            db.session.add_all([
                Result(participant_id=fast.id, project_id=sample_project.id,
                       time_seconds=10, final_time=10),
                Result(participant_id=fast.id, project_id=score_project.id,
                       time_seconds=100, final_time=100),
                Result(participant_id=slow.id, project_id=sample_project.id,
                       time_seconds=20, final_time=20),
                Result(participant_id=slow.id, project_id=score_project.id,
                       time_seconds=1, final_time=1),
            ])
            db.session.commit()

        resp = auth_client.get(f'/admin/activity/{sample_activity.id}/export')

        workbook = load_workbook(io.BytesIO(resp.data))
        sheet = workbook.active
        assert sheet.cell(row=2, column=2).value == 'Fast High Score'
        assert sheet.cell(row=3, column=2).value == 'Slow Low Score'
