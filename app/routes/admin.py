from flask import Blueprint, render_template, request, redirect, url_for, jsonify, send_file, session, flash, abort
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import qrcode
import os
import io
import secrets
from .. import db, logger
from ..models import Activity, ActivityType, Admin, Department, Project, ActivityProject, QRCode, Participant, Recorder, ActivityRecorder, Result, ProjectCategory, SystemLog, Tenant
from ..project_types import get_project_type, project_summary_bucket, project_type_choices, project_type_map
from ..scoring import participant_result_summary, ranking_sort_key
from ..utils import (admin_required, csrf_required, current_tenant_id, format_time, get_current_admin,
                     get_current_tenant, get_or_404, log_action, paginate_query,
                     paginate_request_query,
                     platform_required, safe_float, safe_int, tenant_get_or_404, tenant_required, truncate_str,
                     save_uploaded_file, get_activity_projects)

admin_bp = Blueprint('admin', __name__)


def tenant_id_required():
    tenant_id = current_tenant_id()
    if tenant_id is None:
        abort(403)
    return tenant_id


def scoped(model):
    return model.query.filter_by(tenant_id=tenant_id_required())


def copy_templates_to_tenant(tenant):
    category_map = {}
    for cat in ProjectCategory.query.filter_by(is_template=True, tenant_id=None).order_by(ProjectCategory.sort_order, ProjectCategory.name):
        clone = ProjectCategory(tenant_id=tenant.id, name=cat.name,
                                sort_order=cat.sort_order, is_template=False)
        db.session.add(clone)
        db.session.flush()
        category_map[cat.id] = clone.id
    for project in Project.query.filter_by(is_template=True, tenant_id=None).order_by(Project.name):
        db.session.add(Project(
            tenant_id=tenant.id,
            name=project.name,
            category_id=category_map.get(project.category_id),
            type=project.type,
            max_score=project.max_score,
            penalty_per_violation=project.penalty_per_violation,
            rule_file=project.rule_file,
            video_file=project.video_file,
            is_template=False,
        ))
    for at in ActivityType.query.filter_by(is_template=True, tenant_id=None).order_by(ActivityType.sort_order, ActivityType.name):
        clone = ActivityType(tenant_id=tenant.id, name=at.name,
                             sort_order=at.sort_order, fields_config=at.fields_config,
                             is_template=False)
        db.session.add(clone)
    for dept in Department.query.filter_by(is_template=True, tenant_id=None).order_by(Department.sort_order, Department.name):
        db.session.add(Department(tenant_id=tenant.id, name=dept.name,
                                  sort_order=dept.sort_order, is_template=False))


@admin_bp.before_request
def require_active_tenant_for_admin_routes():
    if request.endpoint and request.endpoint.startswith('admin.') and request.path.startswith('/admin'):
        admin = get_current_admin()
        if not admin:
            return None
        if admin.is_platform_admin and not get_current_tenant():
            return redirect(url_for('admin.platform_tenants'))


@admin_bp.route('/platform')
@platform_required
def platform_index():
    return redirect(url_for('admin.platform_tenants'))


@admin_bp.route('/platform/tenants')
@platform_required
def platform_tenants():
    tenants = Tenant.query.order_by(Tenant.created_at.desc()).all()
    return render_template('platform_tenants.html', tenants=tenants)


@admin_bp.route('/platform/tenant/create', methods=['POST'])
@platform_required
@csrf_required
def platform_create_tenant():
    from ..security import hash_password
    name = truncate_str(request.form.get('name', ''), 100)
    tenant_type = request.form.get('tenant_type', 'school')
    admin_email = truncate_str(request.form.get('admin_email', ''), 120).lower()
    admin_name = truncate_str(request.form.get('admin_name', ''), 50) or admin_email
    password = request.form.get('password', '')
    if not name or not admin_email or not password:
        flash('租户名称、管理员邮箱和密码不能为空', 'danger')
        return redirect(url_for('admin.platform_tenants'))
    if Tenant.query.filter_by(name=name).first():
        flash('学校/单位名称已存在', 'danger')
        return redirect(url_for('admin.platform_tenants'))
    if Admin.query.filter_by(email=admin_email).first():
        flash('管理员邮箱已存在', 'danger')
        return redirect(url_for('admin.platform_tenants'))
    tenant = Tenant(name=name, tenant_type=tenant_type, is_active=True)
    db.session.add(tenant)
    db.session.flush()
    copy_templates_to_tenant(tenant)
    db.session.add(Admin(email=admin_email, username=admin_name,
                         role='tenant_admin', tenant_id=tenant.id,
                         password_hash=hash_password(password)))
    db.session.commit()
    flash(f'已创建学校/单位「{name}」并复制平台模板', 'success')
    return redirect(url_for('admin.platform_tenants'))


@admin_bp.route('/platform/tenant/<int:tenant_id>/toggle', methods=['POST'])
@platform_required
@csrf_required
def platform_toggle_tenant(tenant_id):
    tenant = get_or_404(Tenant, tenant_id)
    tenant.is_active = not tenant.is_active
    db.session.commit()
    return redirect(url_for('admin.platform_tenants'))


@admin_bp.route('/platform/tenant/<int:tenant_id>/manage')
@platform_required
def platform_manage_tenant(tenant_id):
    tenant = get_or_404(Tenant, tenant_id)
    if not tenant.is_active:
        flash('该学校/单位已停用', 'danger')
        return redirect(url_for('admin.platform_tenants'))
    session['active_tenant_id'] = tenant.id
    return redirect(url_for('admin.admin'))


@admin_bp.route('/platform/tenant/exit')
@platform_required
def platform_exit_tenant():
    session.pop('active_tenant_id', None)
    return redirect(url_for('admin.platform_tenants'))


@admin_bp.route('/platform/templates')
@platform_required
def platform_templates():
    categories = ProjectCategory.query.filter_by(is_template=True, tenant_id=None).order_by(ProjectCategory.sort_order, ProjectCategory.name).all()
    projects = Project.query.filter_by(is_template=True, tenant_id=None).order_by(Project.name).all()
    activity_types = ActivityType.query.filter_by(is_template=True, tenant_id=None).order_by(ActivityType.sort_order, ActivityType.name).all()
    departments = Department.query.filter_by(is_template=True, tenant_id=None).order_by(Department.sort_order, Department.name).all()
    return render_template('platform_templates.html', categories=categories, projects=projects,
                           activity_types=activity_types, departments=departments,
                           project_types=project_type_choices())


@admin_bp.route('/platform/template/category/create', methods=['POST'])
@platform_required
@csrf_required
def platform_create_template_category():
    name = truncate_str(request.form.get('name', ''), 50)
    sort_order = safe_int(request.form.get('sort_order', 0), 0)
    if name and not ProjectCategory.query.filter_by(tenant_id=None, is_template=True, name=name).first():
        db.session.add(ProjectCategory(name=name, sort_order=sort_order, is_template=True))
        db.session.commit()
    return redirect(url_for('admin.platform_templates'))


@admin_bp.route('/platform/template/activity-type/create', methods=['POST'])
@platform_required
@csrf_required
def platform_create_template_activity_type():
    name = truncate_str(request.form.get('name', ''), 50)
    sort_order = safe_int(request.form.get('sort_order', 0), 0)
    if name and not ActivityType.query.filter_by(tenant_id=None, is_template=True, name=name).first():
        at = ActivityType(name=name, sort_order=sort_order, is_template=True)
        at.set_fields(at.get_default_fields())
        db.session.add(at)
        db.session.commit()
    return redirect(url_for('admin.platform_templates'))


@admin_bp.route('/platform/template/department/create', methods=['POST'])
@platform_required
@csrf_required
def platform_create_template_department():
    name = truncate_str(request.form.get('name', ''), 100)
    sort_order = safe_int(request.form.get('sort_order', 0), 0)
    if name and not Department.query.filter_by(tenant_id=None, is_template=True, name=name).first():
        db.session.add(Department(name=name, sort_order=sort_order, is_template=True))
        db.session.commit()
    return redirect(url_for('admin.platform_templates'))


@admin_bp.route('/platform/template/project/create', methods=['POST'])
@platform_required
@csrf_required
def platform_create_template_project():
    name = truncate_str(request.form.get('name', ''), 100)
    if not name or Project.query.filter_by(tenant_id=None, is_template=True, name=name).first():
        return redirect(url_for('admin.platform_templates'))
    category_id = safe_int(request.form.get('category_id', 0), 0) or None
    if category_id:
        category = db.session.get(ProjectCategory, category_id)
        if not category or category.tenant_id is not None or not category.is_template:
            category_id = None
    proj_type = request.form.get('type', 'time')
    if proj_type not in project_type_map():
        proj_type = 'time'
    max_score = None
    if get_project_type(proj_type).summary_bucket == 'score':
        max_score = safe_float(request.form.get('max_score', 0), 0) or None
    db.session.add(Project(name=name, category_id=category_id, type=proj_type,
                           max_score=max_score,
                           penalty_per_violation=safe_float(request.form.get('penalty', 5.0), 5.0),
                           is_template=True))
    db.session.commit()
    return redirect(url_for('admin.platform_templates'))


@admin_bp.route('/admin')
@tenant_required
def admin():
    PER_PAGE = 10
    activity_types = scoped(ActivityType).order_by(ActivityType.sort_order, ActivityType.name).all()
    departments = scoped(Department).order_by(Department.sort_order, Department.name).all()

    # Active activities
    act_search = request.args.get('act_search', '').strip()
    act_type_id = request.args.get('act_type_id', 0, type=int)
    act_dept_id = request.args.get('act_dept_id', 0, type=int)
    act_query = scoped(Activity).filter_by(archived=False)
    if act_search:
        act_query = act_query.filter(Activity.name.like(f'%{act_search}%'))
    if act_type_id:
        act_query = act_query.filter_by(activity_type_id=act_type_id)
    if act_dept_id:
        act_query = act_query.filter_by(department_id=act_dept_id)
    active_page = paginate_request_query(
        act_query.order_by(Activity.created_at.desc()), 'act_page', PER_PAGE
    )

    return render_template('admin.html',
                           active=active_page.items,
                           activity_types=activity_types, departments=departments,
                           act_page=active_page.page, act_search=act_search,
                           act_type_id=act_type_id, act_dept_id=act_dept_id,
                           act_total=active_page.total, act_pages=active_page.pages)


# ---- 项目管理 ----

@admin_bp.route('/admin/archived')
@tenant_required
def archived_activities():
    PER_PAGE = 10
    activity_types = scoped(ActivityType).order_by(ActivityType.sort_order, ActivityType.name).all()
    departments = scoped(Department).order_by(Department.sort_order, Department.name).all()

    arc_search = request.args.get('arc_search', '').strip()
    arc_type_id = request.args.get('arc_type_id', 0, type=int)
    arc_dept_id = request.args.get('arc_dept_id', 0, type=int)
    arc_query = scoped(Activity).filter_by(archived=True)
    if arc_search:
        arc_query = arc_query.filter(Activity.name.like(f'%{arc_search}%'))
    if arc_type_id:
        arc_query = arc_query.filter_by(activity_type_id=arc_type_id)
    if arc_dept_id:
        arc_query = arc_query.filter_by(department_id=arc_dept_id)
    archive_page = paginate_request_query(
        arc_query.order_by(Activity.created_at.desc()), 'arc_page', PER_PAGE
    )

    return render_template('archived.html',
                           archived=archive_page.items,
                           activity_types=activity_types, departments=departments,
                           arc_page=archive_page.page, arc_search=arc_search,
                           arc_type_id=arc_type_id, arc_dept_id=arc_dept_id,
                           arc_total=archive_page.total, arc_pages=archive_page.pages)




@admin_bp.route('/admin/projects')
@tenant_required
def project_list():
    PER_PAGE = 20
    search = request.args.get('search', '').strip()
    category_id = request.args.get('category_id', 0, type=int)
    query = scoped(Project)
    if search:
        query = query.filter(Project.name.like(f'%{search}%'))
    if category_id:
        query = query.filter_by(category_id=category_id)
    page_data = paginate_request_query(
        query.order_by(Project.created_at.desc()), 'page', PER_PAGE
    )
    categories = scoped(ProjectCategory).order_by(ProjectCategory.sort_order, ProjectCategory.name).all()
    return render_template('project_list.html', projects=page_data.items, categories=categories,
                           project_types=project_type_map(),
                           current_category_id=category_id, page=page_data.page,
                           pages=page_data.pages, total=page_data.total,
                           search=search)


@admin_bp.route('/admin/project/create', methods=['GET', 'POST'])
@tenant_required
@csrf_required
def create_project():
    tenant_id = tenant_id_required()
    if request.method == 'POST':
        name = truncate_str(request.form.get('name', ''), 100)
        penalty = safe_float(request.form.get('penalty', 5.0), 5.0)
        category_id = safe_int(request.form.get('category_id', 0), 0) or None
        if category_id:
            tenant_get_or_404(ProjectCategory, category_id)
        proj_type = request.form.get('type', 'time')
        if proj_type not in project_type_map():
            proj_type = 'time'
        # 重名检查
        existing = scoped(Project).filter(Project.name == name).first()
        if existing:
            flash(f'项目名称「{name}」已存在，请使用其他名称', 'danger')
            categories = scoped(ProjectCategory).order_by(ProjectCategory.sort_order, ProjectCategory.name).all()
            return render_template('project_form.html', project=None, categories=categories,
                                   project_types=project_type_choices())
        max_score = None
        if get_project_type(proj_type).summary_bucket == 'score':
            max_score = safe_float(request.form.get('max_score', 0), 0) or None
        project = Project(tenant_id=tenant_id, name=name, penalty_per_violation=penalty,
                          category_id=category_id, type=proj_type, max_score=max_score)
        db.session.add(project)
        db.session.flush()

        rule_path = save_uploaded_file(request.files.get('rule_file'), 'rules', 'rule', project.id)
        if rule_path:
            project.rule_file = rule_path
        video_path = save_uploaded_file(request.files.get('video_file'), 'videos', 'video', project.id)
        if video_path:
            project.video_file = video_path

        db.session.commit()
        logger.info(f'创建项目: {project.name} (id={project.id})')
        return redirect(url_for('admin.project_list'))
    categories = scoped(ProjectCategory).order_by(ProjectCategory.sort_order, ProjectCategory.name).all()
    return render_template('project_form.html', project=None, categories=categories,
                           project_types=project_type_choices())


@admin_bp.route('/admin/project/<int:project_id>/edit', methods=['GET', 'POST'])
@tenant_required
@csrf_required
def edit_project(project_id):
    project = tenant_get_or_404(Project, project_id)
    if request.method == 'POST':
        project.name = truncate_str(request.form.get('name', project.name), 100)
        project.penalty_per_violation = safe_float(request.form.get('penalty', project.penalty_per_violation), project.penalty_per_violation)
        project.category_id = safe_int(request.form.get('category_id', 0), 0) or None
        if project.category_id:
            tenant_get_or_404(ProjectCategory, project.category_id)
        proj_type = request.form.get('type', project.type)
        if proj_type in project_type_map():
            project.type = proj_type
        project.max_score = None
        if project_summary_bucket(project) == 'score':
            project.max_score = safe_float(request.form.get('max_score', 0), 0) or None

        rule_path = save_uploaded_file(request.files.get('rule_file'), 'rules', 'rule', project.id)
        if rule_path:
            project.rule_file = rule_path
        video_path = save_uploaded_file(request.files.get('video_file'), 'videos', 'video', project.id)
        if video_path:
            project.video_file = video_path

        db.session.commit()
        logger.info(f'编辑项目: {project.name} (id={project.id})')
        return redirect(url_for('admin.project_list'))
    categories = scoped(ProjectCategory).order_by(ProjectCategory.sort_order, ProjectCategory.name).all()
    return render_template('project_form.html', project=project, categories=categories,
                           project_types=project_type_choices())


@admin_bp.route('/admin/project/<int:project_id>/delete', methods=['POST'])
@tenant_required
@csrf_required
def delete_project(project_id):
    project = tenant_get_or_404(Project, project_id)
    ActivityProject.query.filter_by(project_id=project_id).delete()
    Result.query.filter_by(project_id=project_id).delete()
    db.session.delete(project)
    db.session.commit()
    return redirect(url_for('admin.project_list'))


# ---- 分类管理 ----

@admin_bp.route('/admin/categories')
@tenant_required
def categories():
    all_categories = scoped(ProjectCategory).order_by(ProjectCategory.sort_order, ProjectCategory.name).all()
    return render_template('categories.html', categories=all_categories)


@admin_bp.route('/admin/category/create', methods=['POST'])
@tenant_required
@csrf_required
def create_category():
    tenant_id = tenant_id_required()
    name = truncate_str(request.form.get('name', ''), 50)
    sort_order = safe_int(request.form.get('sort_order', 0), 0)
    if not name:
        return '分类名称不能为空', 400
    exists = scoped(ProjectCategory).filter_by(name=name).first()
    if exists:
        return '分类名称已存在', 400
    db.session.add(ProjectCategory(tenant_id=tenant_id, name=name, sort_order=sort_order))
    db.session.commit()
    logger.info(f'创建项目分组: {name}')
    return redirect(url_for('admin.categories'))


@admin_bp.route('/admin/category/<int:category_id>/edit', methods=['POST'])
@tenant_required
@csrf_required
def edit_category(category_id):
    cat = tenant_get_or_404(ProjectCategory, category_id)
    name = truncate_str(request.form.get('name', ''), 50)
    sort_order = safe_int(request.form.get('sort_order', 0), 0)
    if not name:
        return '分类名称不能为空', 400
    exists = scoped(ProjectCategory).filter(ProjectCategory.name == name, ProjectCategory.id != category_id).first()
    if exists:
        return '分类名称已存在', 400
    cat.name = name
    cat.sort_order = sort_order
    db.session.commit()
    logger.info(f'编辑项目分组: id={category_id}, name={name}')
    return redirect(url_for('admin.categories'))


@admin_bp.route('/admin/category/<int:category_id>/delete', methods=['POST'])
@tenant_required
@csrf_required
def delete_category(category_id):
    cat = tenant_get_or_404(ProjectCategory, category_id)
    if cat.projects:
        return '该分类下有项目，无法删除', 400
    db.session.delete(cat)
    db.session.commit()
    logger.info(f'删除项目分组: id={category_id}')
    return redirect(url_for('admin.categories'))


# ---- 活动类型管理 ----

@admin_bp.route('/admin/activity-types')
@tenant_required
def activity_types():
    types = scoped(ActivityType).order_by(ActivityType.sort_order, ActivityType.name).all()
    return render_template('activity_types.html', types=types)


@admin_bp.route('/admin/activity-type/create', methods=['POST'])
@tenant_required
@csrf_required
def create_activity_type():
    tenant_id = tenant_id_required()
    name = truncate_str(request.form.get('name', ''), 50)
    sort_order = safe_int(request.form.get('sort_order', 0), 0)
    if not name:
        return '类型名称不能为空', 400
    if scoped(ActivityType).filter_by(name=name).first():
        return '类型名称已存在', 400
    at = ActivityType(tenant_id=tenant_id, name=name, sort_order=sort_order)
    at.set_fields(at.get_default_fields())
    db.session.add(at)
    db.session.commit()
    return redirect(url_for('admin.activity_types'))


@admin_bp.route('/admin/activity-type/<int:type_id>/edit', methods=['POST'])
@tenant_required
@csrf_required
def edit_activity_type(type_id):
    at = tenant_get_or_404(ActivityType, type_id)
    name = truncate_str(request.form.get('name', ''), 50)
    sort_order = safe_int(request.form.get('sort_order', 0), 0)
    if not name:
        return '类型名称不能为空', 400
    if scoped(ActivityType).filter(ActivityType.name == name, ActivityType.id != type_id).first():
        return '类型名称已存在', 400
    at.name = name
    at.sort_order = sort_order
    db.session.commit()
    return redirect(url_for('admin.activity_types'))


@admin_bp.route('/admin/activity-type/<int:type_id>/delete', methods=['POST'])
@tenant_required
@csrf_required
def delete_activity_type(type_id):
    at = tenant_get_or_404(ActivityType, type_id)
    if at.activities:
        return '该类型下有活动，无法删除', 400
    db.session.delete(at)
    db.session.commit()
    return redirect(url_for('admin.activity_types'))

@admin_bp.route('/admin/activity-type/<int:type_id>/fields', methods=['POST'])
@tenant_required
@csrf_required
def save_activity_type_fields(type_id):
    import json
    at = tenant_get_or_404(ActivityType, type_id)
    fields_json = request.form.get('fields_config', '[]')
    try:
        fields = json.loads(fields_json)
        if not isinstance(fields, list):
            return '字段配置格式错误', 400
        for f in fields:
            if not isinstance(f, dict) or 'key' not in f or 'label' not in f:
                return '字段配置格式错误', 400
            f['key'] = truncate_str(str(f.get('key', '')).strip(), 50)
            f['label'] = truncate_str(str(f.get('label', '')).strip(), 50)
            f['type'] = f.get('type') if f.get('type') in ('text', 'number', 'select') else 'text'
            f['required'] = bool(f.get('required'))
            if not f['key'] or not f['label']:
                return '字段配置格式错误', 400
            if f['type'] == 'select':
                raw_options = f.get('options')
                if not isinstance(raw_options, list):
                    return '字段配置错误：下拉字段必须配置选项', 400
                options = []
                seen = set()
                for raw in raw_options:
                    option = truncate_str(str(raw).strip(), 100)
                    if not option:
                        continue
                    if ',' in option or '，' in option:
                        return '字段配置错误：下拉选项请使用换行分隔，不要使用逗号', 400
                    if option in seen:
                        return '字段配置错误：下拉选项不能重复', 400
                    seen.add(option)
                    options.append(option)
                if not options:
                    return '字段配置错误：下拉字段必须配置选项', 400
                f['options'] = options
            else:
                f.pop('options', None)
        at.set_fields(fields)
        db.session.commit()
        return redirect(url_for('admin.activity_types'))
    except (json.JSONDecodeError, TypeError):
        return '字段配置格式错误', 400




# ---- 部门管理 ----

@admin_bp.route('/admin/departments')
@tenant_required
def departments():
    PER_PAGE = 20
    page_data = paginate_request_query(
        scoped(Department).order_by(Department.sort_order, Department.name),
        'page', PER_PAGE
    )
    return render_template('departments.html', departments=page_data.items,
                           page=page_data.page, pages=page_data.pages,
                           total=page_data.total)


@admin_bp.route('/admin/department/create', methods=['POST'])
@tenant_required
@csrf_required
def create_department():
    tenant_id = tenant_id_required()
    name = truncate_str(request.form.get('name', ''), 100)
    sort_order = safe_int(request.form.get('sort_order', 0), 0)
    if not name:
        flash('部门名称不能为空', 'danger')
        return redirect(url_for('admin.departments'))
    if scoped(Department).filter_by(name=name).first():
        flash(f'部门名称「{name}」已存在', 'danger')
        return redirect(url_for('admin.departments'))
    db.session.add(Department(tenant_id=tenant_id, name=name, sort_order=sort_order))
    db.session.commit()
    flash(f'部门「{name}」创建成功', 'success')
    return redirect(url_for('admin.departments'))


@admin_bp.route('/admin/department/<int:dept_id>/edit', methods=['POST'])
@tenant_required
@csrf_required
def edit_department(dept_id):
    dept = tenant_get_or_404(Department, dept_id)
    name = truncate_str(request.form.get('name', ''), 100)
    sort_order = safe_int(request.form.get('sort_order', 0), 0)
    if not name:
        flash('部门名称不能为空', 'danger')
        return redirect(url_for('admin.departments'))
    if scoped(Department).filter(Department.name == name, Department.id != dept_id).first():
        flash(f'部门名称「{name}」已存在', 'danger')
        return redirect(url_for('admin.departments'))
    dept.name = name
    dept.sort_order = sort_order
    db.session.commit()
    flash(f'部门名称已修改为「{name}」', 'success')
    return redirect(url_for('admin.departments'))


@admin_bp.route('/admin/department/<int:dept_id>/delete', methods=['POST'])
@tenant_required
@csrf_required
def delete_department(dept_id):
    dept = tenant_get_or_404(Department, dept_id)
    if dept.activities:
        return '该部门下有活动，无法删除', 400
    db.session.delete(dept)
    db.session.commit()
    return redirect(url_for('admin.departments'))


# ---- 活动管理 ----

@admin_bp.route('/admin/activity/create', methods=['GET', 'POST'])
@tenant_required
@csrf_required
def create_activity():
    tenant_id = tenant_id_required()
    all_projects = scoped(Project).order_by(Project.name).all()
    activity_types = scoped(ActivityType).order_by(ActivityType.sort_order, ActivityType.name).all()
    departments = scoped(Department).order_by(Department.sort_order, Department.name).all()
    categories = scoped(ProjectCategory).order_by(ProjectCategory.sort_order, ProjectCategory.name).all()
    if request.method == 'POST':
        name = truncate_str(request.form.get('name', ''), 100)
        activity_type_id = safe_int(request.form.get('activity_type_id', 0), 0) or None
        department_id = safe_int(request.form.get('department_id', 0), 0) or None
        selected_ids = request.form.getlist('project_ids')
        if activity_type_id:
            tenant_get_or_404(ActivityType, activity_type_id)
        if department_id:
            tenant_get_or_404(Department, department_id)

        if scoped(Activity).filter_by(name=name).first():
            return render_template('create_activity.html', all_projects=all_projects,
                                   activity_types=activity_types, departments=departments,
                                   categories=categories,
                                   error='活动名称已存在，请使用其他名称')

        need_class = request.form.get('need_class') == 'on'
        activity = Activity(tenant_id=tenant_id, name=name, activity_type_id=activity_type_id, department_id=department_id, need_class=need_class)
        db.session.add(activity)
        db.session.flush()

        for pid in selected_ids:
            project = tenant_get_or_404(Project, int(pid))
            db.session.add(ActivityProject(activity_id=activity.id, project_id=project.id))

        db.session.commit()
        return redirect(url_for('admin.admin'))
    return render_template('create_activity.html', all_projects=all_projects,
                           activity_types=activity_types, departments=departments,
                           categories=categories)


@admin_bp.route('/admin/activity/<int:activity_id>')
@tenant_required
def activity_detail(activity_id):
    activity = tenant_get_or_404(Activity, activity_id)
    projects = get_activity_projects(activity_id)
    activity_recorders = ActivityRecorder.query.filter_by(activity_id=activity_id).all()
    all_recorders = scoped(Recorder).order_by(Recorder.name).all()
    all_projects = scoped(Project).order_by(Project.name).all()
    categories = scoped(ProjectCategory).order_by(ProjectCategory.sort_order, ProjectCategory.name).all()

    qr_search = request.args.get('qr_search', '').strip()
    qr_name_search = request.args.get('qr_name_search', '').strip()
    qr_query = scoped(QRCode).filter_by(activity_id=activity_id)
    if qr_search:
        qr_query = qr_query.filter(QRCode.code.like(f'%{qr_search}%'))
    if qr_name_search:
        qr_query = qr_query.join(QRCode.participant).filter(
            Participant.name.like(f'%{qr_name_search}%'))
    qr_page_data = paginate_request_query(
        qr_query.order_by(QRCode.id), 'qr_page', 15
    )
    all_qrcodes = scoped(QRCode).filter_by(activity_id=activity_id).all()
    used_count = sum(1 for q in all_qrcodes if q.status == 'used')

    p_search = request.args.get('p_search', '').strip()
    from sqlalchemy import or_
    p_query = scoped(Participant).filter_by(activity_id=activity_id)
    if p_search:
        p_query = p_query.filter(
            or_(Participant.name.like(f'%{p_search}%'),
                Participant.class_name.like(f'%{p_search}%')))
    participant_page = paginate_request_query(
        p_query.order_by(Participant.id), 'p_page', 15
    )
    participants = participant_page.items

    participant_data = []
    results_by_pid = {}
    if participants:
        pid_list = [p.id for p in participants]
        all_results = scoped(Result).filter(Result.participant_id.in_(pid_list)).all()
        for r in all_results:
            results_by_pid.setdefault(r.participant_id, []).append(r)
    for p in participants:
        results = results_by_pid.get(p.id, [])
        participant_data.append(participant_result_summary(p, projects, results))

    return render_template('activity_detail.html',
                           activity=activity, projects=projects, all_projects=all_projects,
                           categories=categories,
                           qrcodes=qr_page_data.items, qr_total=qr_page_data.total,
                           qr_pages=qr_page_data.pages,
                           qr_page=qr_page_data.page, qr_search=qr_search, qr_name_search=qr_name_search,
                           all_qrcodes=all_qrcodes, used_count=used_count,
                           participant_data=participant_data, p_total=participant_page.total,
                           p_pages=participant_page.pages, p_page=participant_page.page, p_search=p_search,
                           activity_recorders=activity_recorders, all_recorders=all_recorders,
                           unused_qrcodes=scoped(QRCode).filter_by(activity_id=activity_id, status='unused').order_by(QRCode.id).all(),
                           all_score=bool(projects) and all(project_summary_bucket(p) == 'score' for p in projects),
                           has_score=any(project_summary_bucket(p) == 'score' for p in projects),
                           has_time=any(project_summary_bucket(p) == 'time' for p in projects),
                           registration_fields=activity.registration_fields)


@admin_bp.route('/admin/activity/<int:activity_id>/project/add', methods=['POST'])
@tenant_required
@csrf_required
def activity_add_project(activity_id):
    activity = tenant_get_or_404(Activity, activity_id)
    if activity.archived:
        return '活动已归档，无法操作', 403
    project_id = safe_int(request.form.get('project_id', 0), 0)
    tenant_get_or_404(Project, project_id)
    exists = ActivityProject.query.filter_by(activity_id=activity_id, project_id=project_id).first()
    if not exists and project_id:
        db.session.add(ActivityProject(activity_id=activity_id, project_id=project_id))
        db.session.commit()
    return redirect(url_for('admin.activity_detail', activity_id=activity_id))


@admin_bp.route('/admin/activity/<int:activity_id>/project/<int:project_id>/remove', methods=['POST'])
@tenant_required
@csrf_required
def activity_remove_project(activity_id, project_id):
    activity = tenant_get_or_404(Activity, activity_id)
    if activity.archived:
        return '活动已归档，无法操作', 403
    ap = ActivityProject.query.filter_by(activity_id=activity_id, project_id=project_id).first()
    if ap:
        db.session.delete(ap)
        db.session.commit()
    return redirect(url_for('admin.activity_detail', activity_id=activity_id))


# ---- 二维码 ----

@admin_bp.route('/admin/activity/<int:activity_id>/qrcode/generate', methods=['POST'])
@tenant_required
@csrf_required
def generate_qrcodes(activity_id):
    tenant_id = tenant_id_required()
    activity = tenant_get_or_404(Activity, activity_id)
    if activity.archived:
        return jsonify({'success': False, 'error': '活动已归档，无法生成二维码'}), 403
    count = safe_int(request.form.get('count', 10), 10)
    import secrets
    # 必须配置自定义词库才能生成
    if not activity.custom_words:
        return jsonify({'success': False, 'error': '请先在⚙️自定义词库中添加词后再生成二维码'}), 400
    WORDS = [w.strip() for w in activity.custom_words.split('\n') if w.strip()]
    if len(WORDS) < 2:
        return jsonify({'success': False, 'error': '自定义词库至少需要 2 个词'}), 400
    qrcodes_data = []
    for _ in range(count):
        for retry in range(100):
            word = secrets.choice(WORDS)
            num = secrets.randbelow(10000)
            code = f'{word}-{num:04d}'
            if not scoped(QRCode).filter_by(code=code).first():
                break
        qr = QRCode(tenant_id=tenant_id, code=code, activity_id=activity_id)
        db.session.add(qr)
        db.session.flush()
        register_url = request.host_url + f'register/{code}'
        qr_img = qrcode.QRCode(version=1, box_size=3, border=2)
        qr_img.add_data(register_url)
        qr_img.make(fit=True)
        img = qr_img.make_image(fill_color='black', back_color='white')
        from flask import current_app
        qr_save_path = os.path.join(current_app.root_path, 'static', 'qrcodes', f'{qr.id}.png')
        img.save(qr_save_path, optimize=True)
        qrcodes_data.append({'code': code, 'url': register_url, 'image': f'qrcodes/{qr.id}.png'})
    db.session.commit()
    log_action('generate_qrcode', detail=f'activity={activity.name}, count={count}', activity_id=activity.id)
    return jsonify({'qrcodes': qrcodes_data})


@admin_bp.route('/admin/activity/<int:activity_id>/qrcode/words', methods=['POST'])
@tenant_required
@csrf_required
def save_qrcode_words(activity_id):
    """保存活动自定义二维码词库"""
    activity = tenant_get_or_404(Activity, activity_id)
    if activity.archived:
        return jsonify({'success': False, 'error': '活动已归档'}), 403
    words = request.form.get('words', '').strip()
    if words:
        lines = [w.strip() for w in words.split('\n') if w.strip()]
        # 去重（保留首次出现顺序）
        seen = set()
        unique = [w for w in lines if not (w in seen or seen.add(w))]
        if len(unique) < 2:
            return jsonify({'success': False, 'error': '至少需要 2 个不同的词'}), 400
        activity.custom_words = '\n'.join(unique)
    else:
        activity.custom_words = None
    db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/admin/activity/<int:activity_id>/qrprint')
@tenant_required
def qr_print(activity_id):
    activity = tenant_get_or_404(Activity, activity_id)
    show_all = request.args.get('all', '0') == '1'
    if show_all:
        qrcodes = scoped(QRCode).filter_by(activity_id=activity_id).order_by(QRCode.status, QRCode.id).all()
    else:
        qrcodes = scoped(QRCode).filter_by(activity_id=activity_id, status='unused').order_by(QRCode.id).all()
    qr_list = []
    for qr in qrcodes:
        item = {'qr': qr, 'participant': None}
        if qr.status == 'used':
            p = scoped(Participant).filter_by(qrcode_id=qr.id).first()
            if p:
                item['participant'] = p
        qr_list.append(item)
    return render_template('qr_print.html', activity=activity, qr_list=qr_list, show_all=show_all)


@admin_bp.route('/admin/activity/<int:activity_id>/qrpdf')
@tenant_required
def qr_pdf(activity_id):
    activity = tenant_get_or_404(Activity, activity_id)
    show_all = request.args.get('all', '0') == '1'
    if show_all:
        qrcodes = scoped(QRCode).filter_by(activity_id=activity_id).order_by(QRCode.status, QRCode.id).all()
    else:
        qrcodes = scoped(QRCode).filter_by(activity_id=activity_id, status='unused').order_by(QRCode.id).all()
    qr_list = []
    for qr in qrcodes:
        item = {'qr': qr, 'participant': None}
        if qr.status == 'used':
            p = scoped(Participant).filter_by(qrcode_id=qr.id).first()
            if p:
                item['participant'] = p
        qr_list.append(item)
    return render_template('qr_print_pdf.html', activity=activity, qr_list=qr_list, show_all=show_all)


# ---- 录入员管理（全局） ----

@admin_bp.route('/admin/recorders')
@tenant_required
def recorder_list():
    PER_PAGE = 20
    q = scoped(Recorder).order_by(Recorder.created_at.desc())
    page_data = paginate_request_query(q, 'page', PER_PAGE)
    # 统计每个录入员分配到的活动数
    ar_counts = db.session.query(
        ActivityRecorder.recorder_id, db.func.count(ActivityRecorder.id)
    ).join(Recorder, Recorder.id == ActivityRecorder.recorder_id).filter(
        Recorder.tenant_id == tenant_id_required()
    ).group_by(ActivityRecorder.recorder_id).all()
    activity_count_map = {r_id: cnt for r_id, cnt in ar_counts}
    return render_template('recorder_list.html', recorders=page_data.items, activity_count_map=activity_count_map,
                           page=page_data.page, pages=page_data.pages, total=page_data.total)


@admin_bp.route('/admin/recorder/create', methods=['POST'])
@tenant_required
@csrf_required
def create_recorder():
    tenant_id = tenant_id_required()
    name = truncate_str(request.form.get('name', ''), 50)
    if not name:
        return '姓名不能为空', 400
    if scoped(Recorder).filter_by(name=name).first():
        flash(f'录入员姓名「{name}」已存在', 'danger')
        return redirect(url_for('admin.recorder_list'))
    import random
    for _ in range(100):
        record_key = f'{random.randint(0, 9999):04d}'
        if not Recorder.query.filter_by(record_key=record_key).first():
            break
    else:
        record_key = f'{random.randint(0, 9999):04d}'
    recorder = Recorder(tenant_id=tenant_id, name=name, record_key=record_key)
    db.session.add(recorder)
    db.session.commit()
    logger.info(f'创建录入员: {name} (key={record_key})')
    flash(f'录入员「{name}」创建成功，KEY: {record_key}', 'success')
    return redirect(url_for('admin.recorder_list'))


@admin_bp.route('/admin/recorder/<int:recorder_id>/edit', methods=['POST'])
@tenant_required
@csrf_required
def edit_recorder(recorder_id):
    recorder = tenant_get_or_404(Recorder, recorder_id)
    name = truncate_str(request.form.get('name', ''), 50)
    if name and name != recorder.name:
        if scoped(Recorder).filter(Recorder.name == name, Recorder.id != recorder_id).first():
            flash(f'录入员姓名「{name}」已存在', 'danger')
            return redirect(url_for('admin.recorder_list'))
        recorder.name = name
    new_key = request.form.get('record_key', '').strip()
    if new_key and len(new_key) == 4 and new_key.isdigit():
        # 检查KEY唯一
        exists = scoped(Recorder).filter(Recorder.record_key == new_key, Recorder.id != recorder_id).first()
        if not exists:
            recorder.record_key = new_key
    db.session.commit()
    logger.info(f'编辑录入员: id={recorder_id}, name={recorder.name}')
    flash(f'录入员「{recorder.name}」已更新', 'success')
    return redirect(url_for('admin.recorder_list'))


@admin_bp.route('/admin/recorder/<int:recorder_id>/delete', methods=['POST'])
@tenant_required
@csrf_required
def delete_recorder(recorder_id):
    recorder = tenant_get_or_404(Recorder, recorder_id)
    ActivityRecorder.query.filter_by(recorder_id=recorder_id).delete()
    Result.query.filter_by(recorder_id=recorder_id).update(
        {'recorder_id': None}
    )
    db.session.delete(recorder)
    db.session.commit()
    logger.info(f'删除录入员: id={recorder_id}')
    return redirect(url_for('admin.recorder_list'))


# ---- 活动-录入员分配 ----

@admin_bp.route('/admin/activity/<int:activity_id>/recorder/assign', methods=['POST'])
@tenant_required
@csrf_required
def assign_recorder_to_activity(activity_id):
    activity = tenant_get_or_404(Activity, activity_id)
    if activity.archived:
        return '活动已归档，无法操作', 403
    recorder_id = safe_int(request.form.get('recorder_id', 0), 0)
    project_ids = ','.join(request.form.getlist('project_ids'))
    if not recorder_id:
        return '请选择录入员', 400
    tenant_get_or_404(Recorder, recorder_id)
    for pid in request.form.getlist('project_ids'):
        tenant_get_or_404(Project, int(pid))
    exists = ActivityRecorder.query.filter_by(activity_id=activity_id, recorder_id=recorder_id).first()
    if exists:
        exists.project_ids = project_ids
    else:
        db.session.add(ActivityRecorder(activity_id=activity_id, recorder_id=recorder_id, project_ids=project_ids))
    db.session.commit()
    logger.info(f'分配录入员: activity={activity_id}, recorder={recorder_id}')
    return redirect(url_for('admin.activity_detail', activity_id=activity_id) + '#tab-recorders')


@admin_bp.route('/admin/activity-recorder/<int:ar_id>/update-projects', methods=['POST'])
@tenant_required
@csrf_required
def update_activity_recorder_projects(ar_id):
    ar = get_or_404(ActivityRecorder, ar_id)
    activity = db.session.get(Activity, ar.activity_id)
    if not activity or activity.tenant_id != tenant_id_required():
        abort(404)
    if activity and activity.archived:
        return '活动已归档，无法操作', 403
    for pid in request.form.getlist('project_ids'):
        tenant_get_or_404(Project, int(pid))
    ar.project_ids = ','.join(request.form.getlist('project_ids'))
    db.session.commit()
    logger.info(f'更新活动录入员项目权限: ar_id={ar_id}')
    return redirect(url_for('admin.activity_detail', activity_id=ar.activity_id) + '#tab-recorders')


@admin_bp.route('/admin/activity-recorder/<int:ar_id>/remove', methods=['POST'])
@tenant_required
@csrf_required
def remove_activity_recorder(ar_id):
    ar = get_or_404(ActivityRecorder, ar_id)
    activity = db.session.get(Activity, ar.activity_id)
    if not activity or activity.tenant_id != tenant_id_required():
        abort(404)
    if activity and activity.archived:
        return '活动已归档，无法操作', 403
    activity_id = ar.activity_id
    db.session.delete(ar)
    db.session.commit()
    logger.info(f'从活动移除录入员: ar_id={ar_id}')
    return redirect(url_for('admin.activity_detail', activity_id=activity_id) + '#tab-recorders')


# ---- 参与者管理 ----

@admin_bp.route('/admin/participant/<int:participant_id>/delete', methods=['POST'])
@tenant_required
@csrf_required
def delete_participant(participant_id):
    participant = tenant_get_or_404(Participant, participant_id)
    activity = db.session.get(Activity, participant.activity_id)
    if activity and activity.archived:
        return '活动已归档，无法操作', 403
    activity_id = participant.activity_id
    qrcode = db.session.get(QRCode, participant.qrcode_id)
    if qrcode:
        qrcode.status = 'unused'
    Result.query.filter_by(participant_id=participant_id).delete()
    log_action('delete_participant', detail=f'name={participant.name}', participant_id=participant_id, activity_id=activity_id)
    db.session.delete(participant)
    db.session.commit()
    return redirect(url_for('admin.activity_detail', activity_id=activity_id))



# ---- 参赛者管理（管理员后台预注册） ----
@admin_bp.route('/admin/activity/<int:activity_id>/participant/add', methods=['POST'])
@tenant_required
@csrf_required
def admin_add_participant(activity_id):
    activity = tenant_get_or_404(Activity, activity_id)
    if activity.archived:
        return '活动已归档', 403
    name = request.form.get('name', '').strip()
    class_name = request.form.get('class_name', '').strip()
    qrcode_id = safe_int(request.form.get('qrcode_id', 0))
    if not name:
        flash('请输入姓名')
        return redirect(url_for('admin.activity_detail', activity_id=activity_id, tab='tab-participants'))
    qr = db.session.get(QRCode, qrcode_id)
    if not qr or qr.tenant_id != activity.tenant_id or qr.activity_id != activity_id or qr.status != 'unused':
        flash('二维码无效或已被使用')
        return redirect(url_for('admin.activity_detail', activity_id=activity_id, tab='tab-participants'))
    participant = Participant(tenant_id=activity.tenant_id, name=name, class_name=class_name,
                              activity_id=activity_id, qrcode_id=qr.id)
    qr.status = 'used'
    try:
        db.session.add(participant)
        db.session.commit()
        log_action('admin_add_participant', detail=f'name={name}, qr={qr.code}',
                   activity_id=activity_id, participant_id=participant.id)
        flash(f'已添加参赛者: {name}')
    except IntegrityError:
        db.session.rollback()
        flash('添加失败，请重试')
    return redirect(url_for('admin.activity_detail', activity_id=activity_id, tab='tab-participants'))


@admin_bp.route('/admin/activity/<int:activity_id>/participant/batch', methods=['POST'])
@tenant_required
@csrf_required
def admin_batch_import_participants(activity_id):
    activity = tenant_get_or_404(Activity, activity_id)
    if activity.archived:
        return '活动已归档', 403
    text = request.form.get('participants_text', '').strip()
    if not text:
        flash('请粘贴参与者名单')
        return redirect(url_for('admin.activity_detail', activity_id=activity_id, tab='tab-participants'))
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    # 获取该活动下所有未使用的二维码
    unused_qrs = scoped(QRCode).filter_by(activity_id=activity_id, status='unused').order_by(QRCode.id).all()
    if len(lines) > len(unused_qrs):
        flash(f'可用二维码不足：需要 {len(lines)} 个，剩余 {len(unused_qrs)} 个')
        return redirect(url_for('admin.activity_detail', activity_id=activity_id, tab='tab-participants'))
    imported = 0
    errors = []
    for i, line in enumerate(lines):
        parts = [p.strip() for p in line.split(',')]
        name = parts[0]
        class_name = parts[1] if len(parts) > 1 else ''
        if not name:
            errors.append(f'第 {i+1} 行: 姓名为空')
            continue
        qr = unused_qrs[imported]
        participant = Participant(tenant_id=activity.tenant_id, name=name, class_name=class_name,
                                  activity_id=activity_id, qrcode_id=qr.id)
        qr.status = 'used'
        try:
            db.session.add(participant)
            db.session.flush()
            log_action('admin_batch_import', detail=f'name={name}, qr={qr.code}',
                       activity_id=activity_id, participant_id=participant.id)
            imported += 1
        except IntegrityError:
            db.session.rollback()
            errors.append(f'第 {i+1} 行 {name}: 导入失败')
    db.session.commit()
    if imported:
        flash(f'成功导入 {imported} 名参赛者')
    for e in errors:
        flash(e)
    return redirect(url_for('admin.activity_detail', activity_id=activity_id, tab='tab-participants'))


# ---- 排名 ----

@admin_bp.route('/admin/activity/<int:activity_id>/ranking')
@tenant_required
def activity_ranking(activity_id):
    activity = tenant_get_or_404(Activity, activity_id)
    projects = get_activity_projects(activity_id)
    has_score = any(project_summary_bucket(p) == 'score' for p in projects)
    has_time = any(project_summary_bucket(p) == 'time' for p in projects)
    all_score = bool(projects) and all(project_summary_bucket(p) == 'score' for p in projects)
    participants = scoped(Participant).filter_by(activity_id=activity_id).all()
    rankings = []
    results_by_pid = {}
    if participants:
        pid_list = [p.id for p in participants]
        all_results = scoped(Result).filter(Result.participant_id.in_(pid_list)).all()
        for r in all_results:
            results_by_pid.setdefault(r.participant_id, []).append(r)
    for p in participants:
        results = results_by_pid.get(p.id, [])
        pr = {}
        for proj in projects:
            found = [r for r in results if r.project_id == proj.id]
            pr[proj.id] = found[0].final_time if found else None
        summary = participant_result_summary(p, projects, results)
        summary['project_results'] = pr
        rankings.append(summary)
    # 时间型升序（越小越好），分数型降序（越大越好）
    rankings.sort(key=ranking_sort_key)
    best_score_total = max((d['score_total'] for d in rankings), default=None)
    project_rankings = {}
    for proj in projects:
        pr_list = [d for d in rankings if d['project_results'].get(proj.id) is not None]
        if project_summary_bucket(proj) == 'score':
            pr_list.sort(key=lambda x: -(x['project_results'][proj.id]))
        else:
            pr_list.sort(key=lambda x: x['project_results'][proj.id])
        project_rankings[proj.id] = pr_list
    return render_template('ranking.html', activity=activity, projects=projects,
                           rankings=rankings, project_rankings=project_rankings,
                           all_score=all_score, has_score=has_score,
                           has_time=has_time,
                           best_score_total=best_score_total,
                           registration_fields=activity.registration_fields)


# ---- 导出 ----

@admin_bp.route('/admin/activity/<int:activity_id>/export')
@tenant_required
def export_results(activity_id):
    activity = tenant_get_or_404(Activity, activity_id)
    projects = get_activity_projects(activity_id)
    has_score = any(project_summary_bucket(p) == 'score' for p in projects)
    all_score = bool(projects) and all(project_summary_bucket(p) == 'score' for p in projects)
    has_time = any(project_summary_bucket(p) == 'time' for p in projects)
    participants = scoped(Participant).filter_by(activity_id=activity_id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = activity.name
    headers = ['排名', '姓名']
    if activity.is_student:
        headers.append('班级')
    for f in activity.registration_fields:
        if f['key'] not in ('name', 'class_name'):
            headers.append(f['label'])
    for proj in projects:
        headers.extend(get_project_type(proj.type).export_headers(proj))
    headers.append('总成绩' if all_score else '总用时')
    if has_score and has_time:
        headers.append('总成绩')
    rankings = []
    results_by_pid = {}
    if participants:
        pid_list = [p.id for p in participants]
        all_results = scoped(Result).filter(Result.participant_id.in_(pid_list)).all()
        for r in all_results:
            results_by_pid.setdefault(r.participant_id, []).append(r)
    for p in participants:
        results = results_by_pid.get(p.id, [])
        rankings.append(participant_result_summary(p, projects, results))
    rankings.sort(key=ranking_sort_key)
    for col, header in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = Font(bold=True); c.alignment = Alignment(horizontal='center')
    for rank, data in enumerate(rankings, 1):
        row, col = rank + 1, 1
        ws.cell(row=row, column=col, value=rank); col += 1
        ws.cell(row=row, column=col, value=data['participant'].name); col += 1
        if activity.is_student:
            ws.cell(row=row, column=col, value=data['participant'].class_name or ''); col += 1
        for f in activity.registration_fields:
            if f['key'] not in ('name', 'class_name'):
                ws.cell(row=row, column=col, value=data['extra'].get(f['key'], '')); col += 1
        for proj in projects:
            p_results = results_by_pid.get(data['participant'].id, [])
            r = next((r for r in p_results if r.project_id == proj.id), None)
            for value in get_project_type(proj.type).export_values(proj, r):
                ws.cell(row=row, column=col, value=value)
                col += 1
        if all_score:
            ws.cell(row=row, column=col, value=data['score_total'])
        else:
            ws.cell(row=row, column=col, value=format_time(data['time_total']))
        if has_score and has_time:
            col += 1
            ws.cell(row=row, column=col, value=data['score_total'])
    for col in range(1, len(headers) + 1):
        letter = ''
        c = col
        while c > 0:
            c, r = divmod(c - 1, 26)
            letter = chr(65 + r) + letter
        ws.column_dimensions[letter].width = 20
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{activity.name}_成绩.xlsx')


# ---- 删除活动 ----

@admin_bp.route('/admin/activity/<int:activity_id>/archive', methods=['POST'])
@tenant_required
@csrf_required
def archive_activity(activity_id):
    activity = tenant_get_or_404(Activity, activity_id)
    activity.archived = True
    db.session.commit()
    logger.info(f'归档活动: {activity.name} (id={activity_id})')
    return redirect(url_for('admin.admin'))


@admin_bp.route('/admin/activity/<int:activity_id>/unarchive', methods=['POST'])
@tenant_required
@csrf_required
def unarchive_activity(activity_id):
    activity = tenant_get_or_404(Activity, activity_id)
    activity.archived = False
    db.session.commit()
    logger.info(f'恢复活动: {activity.name} (id={activity_id})')
    return redirect(url_for('admin.admin'))


@admin_bp.route('/admin/activity/<int:activity_id>/delete', methods=['POST'])
@tenant_required
@csrf_required
def delete_activity(activity_id):
    activity = tenant_get_or_404(Activity, activity_id)
    for p in scoped(Participant).filter_by(activity_id=activity_id).all():
        Result.query.filter_by(participant_id=p.id).delete()
    scoped(Participant).filter_by(activity_id=activity_id).delete()
    from flask import current_app
    for qr in scoped(QRCode).filter_by(activity_id=activity_id).all():
        qr_path = os.path.join(current_app.root_path, 'static', 'qrcodes', f'{qr.id}.png')
        if os.path.exists(qr_path):
            os.remove(qr_path)
    scoped(QRCode).filter_by(activity_id=activity_id).delete()
    ActivityRecorder.query.filter_by(activity_id=activity_id).delete()
    ActivityProject.query.filter_by(activity_id=activity_id).delete()
    db.session.delete(activity)
    db.session.commit()
    logger.info(f'删除活动: activity_id={activity_id}')
    return redirect(url_for('admin.admin'))

# ---- 系统日志 ----
@admin_bp.route('/admin/logs')
@tenant_required
def system_logs():
    page = request.args.get('page', 1, type=int)
    action_filter = request.args.get('action', '')
    from app.models import SystemLog
    query = scoped(SystemLog)
    if action_filter:
        query = query.filter(SystemLog.action == action_filter)
    query = query.order_by(SystemLog.created_at.desc())
    items, total, pages = paginate_query(query, page, 50)
    actions = db.session.query(SystemLog.action).filter(
        SystemLog.tenant_id == tenant_id_required()
    ).distinct().order_by(SystemLog.action).all()
    actions = [a[0] for a in actions]
    return render_template('admin_logs.html', logs=items, page=page, pages=pages, total=total,
                           actions=actions, action_filter=action_filter)


# ---- 参赛者历史记录 ----
@admin_bp.route('/admin/participant/<int:participant_id>/history')
@tenant_required
def participant_history(participant_id):
    participant = tenant_get_or_404(Participant, participant_id)
    activity = tenant_get_or_404(Activity, participant.activity_id)
    results = scoped(Result).filter_by(participant_id=participant_id).order_by(Result.recorded_at.desc()).all()
    result_list = []
    for r in results:
        project = db.session.get(Project, r.project_id)
        recorder = db.session.get(Recorder, r.recorder_id) if r.recorder_id else None
        result_list.append({
            'project': project.name if project else f'(#{r.project_id})',
            'project_type': project.type if project else 'time',
            'value': r.final_time,
            'violations': r.violations,
            'recorded_at': r.recorded_at,
            'recorder_name': recorder.name if recorder else '-',
        })
    return render_template('participant_history.html', participant=participant, activity=activity,
                           results=result_list)
